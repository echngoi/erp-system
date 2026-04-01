"""
API Views for ZKTeco Attendance System
"""
import io
import logging
from collections import defaultdict
from datetime import datetime, timedelta, time as dtime
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Count, Q, Min, Max
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view

from .models import Employee, AttendanceLog, SyncLog, AttendancePermission, WorkShift, LateEarlyRule, PenaltyConfig
from .serializers import EmployeeSerializer, AttendanceLogSerializer, SyncLogSerializer, AttendancePermissionSerializer, WorkShiftSerializer, LateEarlyRuleSerializer, PenaltyConfigSerializer
from .permissions import get_allowed_pages, can_view_all_on_page
from .zk_service import zk_service
from users.permissions import IsAdmin

logger = logging.getLogger(__name__)

ALLOWED_DEVICE_COMMANDS = {'clearlog', 'reboot', 'getuser', 'getlog'}


class DeviceStatusView(APIView):
    """Test connection to device."""
    permission_classes = [IsAdmin]

    def get(self, request):
        result = zk_service.test_connection()
        return Response(result)


class DeviceTimeView(APIView):
    """Get/set device time."""
    permission_classes = [IsAdmin]

    def get(self, request):
        try:
            result = zk_service.get_device_time()
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        """Sync device time with server."""
        try:
            zk_service.set_device_time()
            return Response({'message': 'Đồng bộ thời gian thành công'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeviceRestartView(APIView):
    """Restart device."""
    permission_classes = [IsAdmin]

    def post(self, request):
        try:
            zk_service.restart_device()
            return Response({'message': 'Đã gửi lệnh khởi động lại thiết bị'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeviceCommandView(APIView):
    """Send command to device via WebSocket command queue."""
    permission_classes = [IsAdmin]

    def post(self, request):
        from .zk_service import adms_enqueue_command, adms_get_last_contact

        cmd = request.data.get('command', '').strip()
        if not cmd:
            return Response({'error': 'Thiếu tham số command'}, status=status.HTTP_400_BAD_REQUEST)
        if cmd not in ALLOWED_DEVICE_COMMANDS:
            return Response(
                {'error': f'Lệnh không hợp lệ: {cmd}. Chấp nhận: {", ".join(sorted(ALLOWED_DEVICE_COMMANDS))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        contact = adms_get_last_contact()
        sn = contact['sn'] if contact else None
        adms_enqueue_command(cmd, sn=sn)

        return Response({
            'message': f'Lệnh "{cmd}" đã được đưa vào hàng đợi. Sẽ gửi khi máy chấm công heartbeat lần tiếp.',
            'command': cmd,
            'target_sn': sn,
        })

    def get(self, request):
        from .zk_service import adms_get_pending_commands, adms_get_command_results
        commands = adms_get_pending_commands()
        results = adms_get_command_results()
        return Response({
            'pending': [
                {'cmd': c['cmd'], 'sn': c['sn'], 'queued_at': c['queued_at'].strftime('%Y-%m-%d %H:%M:%S')}
                for c in commands
            ],
            'results': [
                {'cmd': r['cmd'], 'sn': r['sn'], 'result': r['result'], 'msg': r['msg'],
                 'received_at': r['received_at'].strftime('%Y-%m-%d %H:%M:%S')}
                for r in results
            ],
        })


class SyncUsersView(APIView):
    """Pull users from device and save to DB."""
    permission_classes = [IsAdmin]

    def post(self, request):
        try:
            device_users = zk_service.get_users()
            created = 0
            updated = 0
            for u in device_users:
                obj, is_new = Employee.objects.update_or_create(
                    uid=u['uid'],
                    defaults={
                        'user_id': str(u['user_id']),
                        'name': u['name'] or f"User {u['user_id']}",
                        'privilege': u['privilege'],
                        'group_id': str(u['group_id'] or ''),
                        'card': u['card'] or 0,
                    }
                )
                if is_new:
                    created += 1
                else:
                    updated += 1
            return Response({
                'message': f'Đồng bộ thành công: {created} mới, {updated} cập nhật',
                'total': len(device_users),
                'created': created,
                'updated': updated,
            })
        except Exception as e:
            logger.error(f"Sync users error: {e}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SyncAttendanceView(APIView):
    """Pull attendance records from device."""
    permission_classes = [IsAdmin]

    def post(self, request):
        from .zk_service import ADMSService

        # ADMS mode: dữ liệu đã được push tự động vào DB, không cần sync thủ công
        if isinstance(zk_service, ADMSService):
            total = AttendanceLog.objects.count()
            return Response({
                'message': f'ADMS mode: máy chấm công tự đẩy dữ liệu về server. Hiện có {total} bản ghi.',
                'total_on_device': total,
                'new_records': 0,
                'note': 'Không cần đồng bộ thủ công trong chế độ ADMS. Dữ liệu được push tự động khi nhân viên chấm công.',
            })

        sync_log = SyncLog.objects.create(status='partial')
        try:
            records = zk_service.get_all_attendance()
            new_count = 0
            for rec in records:
                if not rec['timestamp']:
                    continue
                ts = datetime.strptime(rec['timestamp'], '%Y-%m-%d %H:%M:%S')
                ts = timezone.make_aware(ts)
                employee = Employee.objects.filter(user_id=str(rec['user_id'])).first()
                _, created = AttendanceLog.objects.get_or_create(
                    user_id=str(rec['user_id']),
                    timestamp=ts,
                    defaults={
                        'employee': employee,
                        'status': rec['status'],
                        'punch': rec['punch'],
                    }
                )
                if created:
                    new_count += 1

            sync_log.status = 'success'
            sync_log.records_synced = new_count
            sync_log.finished_at = timezone.now()
            sync_log.save()

            return Response({
                'message': f'Đồng bộ thành công: {new_count} bản ghi mới',
                'total_on_device': len(records),
                'new_records': new_count,
            })
        except Exception as e:
            sync_log.status = 'failed'
            sync_log.error_message = str(e)
            sync_log.finished_at = timezone.now()
            sync_log.save()
            logger.error(f"Sync attendance error: {e}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ClearAttendanceView(APIView):
    """Clear all attendance logs on device."""
    permission_classes = [IsAdmin]

    def post(self, request):
        try:
            zk_service.clear_attendance()
            return Response({'message': 'Đã xóa toàn bộ dữ liệu chấm công trên thiết bị'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EmployeeListView(APIView):
    """List all employees (from local DB)."""
    permission_classes = [IsAdmin]

    def get(self, request):
        employees = Employee.objects.select_related('linked_user__department', 'shift').all()
        # By default only return active; admin can pass ?show_all=1
        if not request.query_params.get('show_all'):
            employees = employees.filter(is_active=True)
        serializer = EmployeeSerializer(employees, many=True)
        return Response({'results': serializer.data, 'total': employees.count()})


class AttendanceListView(APIView):
    """List attendance records with filters. Non-admin only sees own data unless granted."""

    def get(self, request):
        user = request.user
        allowed = get_allowed_pages(user)
        if 'logs' not in allowed:
            return Response({'detail': 'Bạn không có quyền truy cập trang này.'}, status=status.HTTP_403_FORBIDDEN)

        qs = AttendanceLog.objects.select_related('employee__linked_user__department')
        if not can_view_all_on_page(user, 'logs'):
            att_emp = getattr(user, 'attendance_employee', None)
            if att_emp:
                qs = qs.filter(user_id=att_emp.user_id)
            else:
                qs = qs.none()

        # Filters
        user_id = request.query_params.get('user_id')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        punch = request.query_params.get('punch')

        if user_id:
            qs = qs.filter(user_id=user_id)
        if punch is not None:
            qs = qs.filter(punch=punch)
        if date_from:
            try:
                df = datetime.strptime(date_from, '%Y-%m-%d')
                qs = qs.filter(timestamp__gte=timezone.make_aware(df))
            except ValueError:
                pass
        if date_to:
            try:
                dt = datetime.strptime(date_to, '%Y-%m-%d')
                dt = dt.replace(hour=23, minute=59, second=59)
                qs = qs.filter(timestamp__lte=timezone.make_aware(dt))
            except ValueError:
                pass

        # Pagination
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 50))
        total = qs.count()
        start = (page - 1) * page_size
        qs = qs[start:start + page_size]

        serializer = AttendanceLogSerializer(qs, many=True)
        return Response({
            'results': serializer.data,
            'total': total,
            'page': page,
            'page_size': page_size,
        })


class DashboardStatsView(APIView):
    """Dashboard statistics."""
    permission_classes = [IsAdmin]

    def get(self, request):
        now_local = timezone.localtime(timezone.now())
        today = now_local.date()
        week_start = today - timedelta(days=today.weekday())
        month_start = today.replace(day=1)

        # Use local-time-aware range filters for correct date boundaries
        from datetime import time as _t
        today_start = timezone.make_aware(datetime.combine(today, _t.min))
        today_end = timezone.make_aware(datetime.combine(today, _t.max))

        today_count = AttendanceLog.objects.filter(
            timestamp__range=(today_start, today_end)
        ).count()

        week_start_dt = timezone.make_aware(datetime.combine(week_start, _t.min))
        week_count = AttendanceLog.objects.filter(
            timestamp__gte=week_start_dt
        ).count()

        month_start_dt = timezone.make_aware(datetime.combine(month_start, _t.min))
        month_count = AttendanceLog.objects.filter(
            timestamp__gte=month_start_dt
        ).count()

        total_employees = Employee.objects.count()

        # Active today (checked in)
        active_today = AttendanceLog.objects.filter(
            timestamp__range=(today_start, today_end),
            punch=0,
        ).values('user_id').distinct().count()

        # Recent 7 days chart
        daily_stats = []
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            d_start = timezone.make_aware(datetime.combine(d, _t.min))
            d_end = timezone.make_aware(datetime.combine(d, _t.max))
            count = AttendanceLog.objects.filter(timestamp__range=(d_start, d_end)).count()
            daily_stats.append({
                'date': d.strftime('%d/%m'),
                'count': count,
            })

        last_sync = SyncLog.objects.filter(status='success').first()

        return Response({
            'today_checkins': today_count,
            'week_checkins': week_count,
            'month_checkins': month_count,
            'total_employees': total_employees,
            'active_today': active_today,
            'daily_stats': daily_stats,
            'last_sync': timezone.localtime(last_sync.started_at).strftime('%Y-%m-%d %H:%M:%S') if last_sync else None,
        })


class SyncLogListView(APIView):
    """List sync history."""
    permission_classes = [IsAdmin]

    def get(self, request):
        logs = SyncLog.objects.all()[:20]
        serializer = SyncLogSerializer(logs, many=True)
        return Response({'results': serializer.data})


class LiveAttendanceView(APIView):
    """Fetch real-time attendance directly from device (no cache)."""
    permission_classes = [IsAdmin]

    def get(self, request):
        try:
            records = zk_service.get_all_attendance()
            records.sort(key=lambda x: x['timestamp'] or '', reverse=True)
            return Response({
                'records': records[:100],
                'total': len(records),
            })
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeviceProtocolView(APIView):
    """Trả về giao thức đang dùng và hướng dẫn cấu hình."""
    permission_classes = [IsAdmin]

    def get(self, request):
        from .zk_service import ADMSService, ZKBinaryService
        proto = 'adms' if isinstance(zk_service, ADMSService) else 'binary'

        # Lấy IP server thực tế từ request
        server_host = request.get_host().split(':')[0]
        server_port = request.get_port() if hasattr(request, 'get_port') else '8000'

        return Response({
            'protocol': proto,
            'device_ip': zk_service.ip,
            'device_port': zk_service.port,
            'server_ip': server_host,
            'adms_push_url': f'http://{server_host}:{server_port}/iclock/cdata',
            'instructions': (
                'ADMS mode: Vào máy chấm công → Menu → Comm. → Cloud Server Setting → '
                f'đặt Domain/IP = {server_host}, Port = {server_port}. '
                'Máy sẽ tự đẩy dữ liệu chấm công về server.'
                if proto == 'adms'
                else 'Binary mode: Django kết nối thẳng ra máy qua pyzk.'
            ),
        })


class ADMSPushView(APIView):
    """
    Endpoint nhận dữ liệu PUSH từ máy chấm công Ronald Jack / ZKTeco ADMS.

    Máy sẽ gửi HTTP GET/POST đến URL này khi:
      - Khởi động
      - Có chấm công mới
      - Theo lịch định kỳ

    Các request máy gửi:
      GET  /api/adms/push/?SN=<serial>&options=...     ← máy xin lệnh
      POST /api/adms/push/?SN=<serial>&table=ATTLOG    ← máy đẩy chấm công
      POST /api/adms/push/?SN=<serial>&table=OPERLOG   ← máy đẩy log vận hành
    """

    authentication_classes = []  # Máy chấm công không gửi auth
    permission_classes     = []

    def get(self, request):
        """Máy gửi GET để xin lệnh. Trả về OK để máy tiếp tục."""
        sn = request.GET.get('SN', 'unknown')
        logger.info(f"ADMS GET từ máy SN={sn}: {dict(request.GET)}")
        # Trả về OK + timestamp server để máy sync giờ
        server_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return Response(
            f"GET OPTION FROM: {sn}\nDatetime={server_time}\n",
            content_type='text/plain'
        )

    def post(self, request):
        """Máy gửi POST với dữ liệu chấm công."""
        sn    = request.GET.get('SN', 'unknown')
        table = request.GET.get('table', '').upper()
        logger.info(f"ADMS POST từ máy SN={sn}, table={table}")

        if table == 'ATTLOG':
            return self._handle_attlog(request, sn)
        elif table == 'OPERLOG':
            logger.info(f"ADMS OPERLOG: {request.body[:200]}")
            return Response("OK", content_type='text/plain')
        else:
            # Một số firmware gửi body dạng ATTLOG trực tiếp
            body = request.body.decode('utf-8', errors='ignore')
            if '\t' in body or 'ATTLOG' in body:
                return self._parse_attlog_body(body, sn)
            logger.info(f"ADMS unknown table={table}, body={body[:100]}")
            return Response("OK", content_type='text/plain')

    def _handle_attlog(self, request, sn):
        """Parse và lưu attendance log từ máy."""
        body = request.body.decode('utf-8', errors='ignore')
        count = self._parse_attlog_body(body, sn)
        return Response(f"OK: {count}", content_type='text/plain')

    def _parse_attlog_body(self, body, sn):
        """
        Format ATTLOG của ZKTeco/Ronald Jack:
          ATTLOG\t<UserID>\t<Timestamp>\t<Status>\t<Punch>\t...\n
        hoặc chỉ:
          <UserID>\t<Timestamp>\t<Status>\t<Punch>\n
        """
        saved = 0
        lines = body.strip().split('\n')
        for line in lines:
            line = line.strip()
            if not line or line.startswith('ATTLOG'):
                # Header line, bỏ qua
                if '\t' in line:
                    parts = line.split('\t')
                    if parts[0] == 'ATTLOG':
                        parts = parts[1:]
                    if len(parts) >= 2:
                        line = '\t'.join(parts)
                    else:
                        continue
                else:
                    continue

            parts = line.split('\t')
            if len(parts) < 2:
                continue

            try:
                user_id   = parts[0].strip()
                ts_str    = parts[1].strip()
                status    = int(parts[2]) if len(parts) > 2 else 0
                punch     = int(parts[3]) if len(parts) > 3 else 0

                # Parse timestamp: "2026-03-28 09:05:44" hoặc "2026/03/28 09:05:44"
                ts_str = ts_str.replace('/', '-')
                ts = datetime.strptime(ts_str, '%Y-%m-%d %H:%M:%S')
                ts = timezone.make_aware(ts)

                employee = Employee.objects.filter(user_id=user_id).first()
                _, created = AttendanceLog.objects.get_or_create(
                    user_id=user_id,
                    timestamp=ts,
                    defaults={'employee': employee, 'status': status, 'punch': punch}
                )
                if created:
                    saved += 1
                    logger.info(f"ADMS: Lưu chấm công user={user_id} ts={ts}")
            except Exception as e:
                logger.warning(f"ADMS parse lỗi dòng '{line}': {e}")

        return saved


# ─── Attendance Report ───────────────────────────────────────────────

# Default work schedule (configurable)
WORK_START = dtime(8, 0)    # 08:00
WORK_END   = dtime(17, 0)   # 17:00


def _match_late_early_rule(rules, rule_type, minutes):
    """Find the matching LateEarlyRule label for given minutes."""
    for r in rules:
        if r.rule_type != rule_type:
            continue
        if minutes >= r.from_minutes:
            if r.to_minutes is None or minutes <= r.to_minutes:
                return r.label or f'{rule_type} {r.from_minutes}-{r.to_minutes or "∞"}p'
    return None


def _calc_penalty(penalty_configs, rule_type, count):
    """Calculate total penalty using tiered calculation.

    Each tier defines a per-occurrence amount for violations falling
    within [from_count, to_count].  Violations beyond all defined tiers
    use the highest tier’s rate.

    Example:
        Tier 1: from=1, to=2, amount=30000  → occurrences 1-2 charged at 30k each
        Tier 2: from=3, to=5, amount=60000  → occurrences 3-5 charged at 60k each
        If count=7: (2×30k) + (3×60k) + (2×60k) = 60k + 180k + 120k = 360k
    """
    if count <= 0:
        return 0

    # Filter and sort tiers for this rule_type
    tiers = sorted(
        [pc for pc in penalty_configs if pc.rule_type == rule_type],
        key=lambda pc: pc.from_count,
    )
    if not tiers:
        return 0

    total = 0
    for tier in tiers:
        if count < tier.from_count:
            break  # count is below this tier, done
        upper = min(count, tier.to_count) if tier.to_count is not None else count
        occurrences = upper - tier.from_count + 1
        total += occurrences * tier.penalty_amount

    # Handle counts beyond all defined tiers (use highest tier’s rate)
    last_tier = tiers[-1]
    if last_tier.to_count is not None and count > last_tier.to_count:
        overflow = count - last_tier.to_count
        total += overflow * last_tier.penalty_amount

    return total


def _detect_missing_punches(shift, punch_times):
    """Detect missing punches based on shift configuration.
    Returns list of missing punch labels.
    """
    if not shift or not punch_times:
        return []

    slots = shift.time_slots  # [(label, time), ...]
    expected = len(slots)
    actual = len(punch_times)

    missing = []
    if actual < expected:
        # First punch missing → no check-in
        if actual == 0:
            missing = [label for label, _ in slots]
        elif shift.shift_type == 'hc':
            # 2 punches expected
            if actual == 1:
                # Only one punch: determine if it's in or out based on proximity
                p = punch_times[0]
                mid = datetime.combine(datetime.today(), shift.start_time) + \
                    (datetime.combine(datetime.today(), shift.end_time) - datetime.combine(datetime.today(), shift.start_time)) / 2
                mid_t = mid.time()
                if p <= mid_t:
                    missing.append('Ra')
                else:
                    missing.append('Vào')
        elif shift.shift_type == '3punch':
            # 3 punches expected: in, mid, out
            if actual == 1:
                missing.extend(['Giữa ca', 'Ra'])
            elif actual == 2:
                missing.append('Giữa ca')
        elif shift.shift_type == '4punch':
            # 4 punches expected
            if actual == 1:
                missing.extend(['Ra ca 1', 'Vào ca 2', 'Ra ca 2'])
            elif actual == 2:
                missing.extend(['Vào ca 2', 'Ra ca 2'])
            elif actual == 3:
                # Determine which one is missing
                missing.append('Giữa ca')  # generic

    return missing


def _build_report_data(date_from, date_to, user_id=None):
    """Build attendance report data for the given date range.

    Returns list of dicts with per-employee daily summary.
    """
    qs = AttendanceLog.objects.select_related('employee__linked_user__department').filter(
        timestamp__date__gte=date_from,
        timestamp__date__lte=date_to,
    )
    if user_id:
        qs = qs.filter(user_id=user_id)

    # Group logs by (user_id, date) — use local time for grouping
    user_day_logs = defaultdict(list)
    for log in qs.order_by('timestamp'):
        local_ts = timezone.localtime(log.timestamp)
        day = local_ts.date()
        user_day_logs[(log.user_id, day)].append(log)

    # Get all employees (or filtered), only active ones for reports
    emp_qs = Employee.objects.select_related('linked_user__department', 'shift').filter(is_active=True)
    if user_id:
        emp_qs = Employee.objects.select_related('linked_user__department', 'shift').filter(user_id=user_id)
    employees = {e.user_id: e for e in emp_qs}

    # Preload late/early rules and penalty configs per shift
    shift_ids = {e.shift_id for e in employees.values() if e.shift_id}
    rules_by_shift = defaultdict(list)
    penalties_by_shift = defaultdict(list)
    if shift_ids:
        for r in LateEarlyRule.objects.filter(shift_id__in=shift_ids).order_by('from_minutes'):
            rules_by_shift[r.shift_id].append(r)
        for p in PenaltyConfig.objects.filter(shift_id__in=shift_ids).order_by('from_count'):
            penalties_by_shift[p.shift_id].append(p)

    # Build date range list
    days = []
    d = date_from
    while d <= date_to:
        days.append(d)
        d += timedelta(days=1)

    rows = []
    for emp_uid, emp in employees.items():
        total_late = 0
        total_early = 0
        total_absent = 0
        total_ot_minutes = 0
        total_work_minutes = 0
        total_late_minutes = 0
        total_early_minutes = 0
        daily = []

        shift = emp.shift
        work_start = shift.start_time if shift else WORK_START
        work_end = shift.end_time if shift else WORK_END
        shift_rules = rules_by_shift.get(shift.id, []) if shift else []
        shift_penalties = penalties_by_shift.get(shift.id, []) if shift else []

        for day in days:
            logs = user_day_logs.get((emp_uid, day), [])
            if not logs:
                total_absent += 1
                # Detect missing punches when employee is absent and has a shift
                mp = [label for label, _ in shift.time_slots] if shift else []
                daily.append({
                    'date': day.strftime('%Y-%m-%d'),
                    'status': 'absent',
                    'check_in': None,
                    'check_out': None,
                    'punches': [],
                    'late_minutes': 0,
                    'early_minutes': 0,
                    'ot_minutes': 0,
                    'work_minutes': 0,
                    'late_label': None,
                    'early_label': None,
                    'missing_punches': mp,
                })
                continue

            # All punches for this day
            punch_times = [timezone.localtime(log.timestamp).time() for log in logs]

            first_ts = timezone.localtime(logs[0].timestamp)
            last_ts = timezone.localtime(logs[-1].timestamp)
            check_in_time = first_ts.time()
            check_out_time = last_ts.time() if len(logs) > 1 else None

            late_min = 0
            early_min = 0
            ot_min = 0
            work_min = 0
            day_status = 'present'
            late_label = None
            early_label = None

            # Late?
            if check_in_time > work_start:
                delta = datetime.combine(day, check_in_time) - datetime.combine(day, work_start)
                late_min = int(delta.total_seconds() / 60)
                if late_min > 0:
                    day_status = 'late'
                    total_late += 1
                    total_late_minutes += late_min
                    late_label = _match_late_early_rule(shift_rules, 'late', late_min)

            # Early leave?
            if check_out_time and check_out_time < work_end:
                delta = datetime.combine(day, work_end) - datetime.combine(day, check_out_time)
                early_min = int(delta.total_seconds() / 60)
                if early_min > 0:
                    if day_status == 'late':
                        day_status = 'late+early'
                    else:
                        day_status = 'early'
                    total_early += 1
                    total_early_minutes += early_min
                    early_label = _match_late_early_rule(shift_rules, 'early', early_min)

            # Overtime?
            if check_out_time and check_out_time > work_end:
                delta = datetime.combine(day, check_out_time) - datetime.combine(day, work_end)
                ot_min = int(delta.total_seconds() / 60)
                total_ot_minutes += ot_min

            # Work duration
            if check_out_time:
                work_delta = datetime.combine(day, check_out_time) - datetime.combine(day, check_in_time)
                work_min = int(work_delta.total_seconds() / 60)
                total_work_minutes += work_min

            # Missing punches detection
            missing_p = _detect_missing_punches(shift, punch_times)

            daily.append({
                'date': day.strftime('%Y-%m-%d'),
                'status': day_status,
                'check_in': check_in_time.strftime('%H:%M:%S'),
                'check_out': check_out_time.strftime('%H:%M:%S') if check_out_time else None,
                'punches': [t.strftime('%H:%M:%S') for t in punch_times],
                'late_minutes': late_min,
                'early_minutes': early_min,
                'ot_minutes': ot_min,
                'work_minutes': work_min,
                'late_label': late_label,
                'early_label': early_label,
                'missing_punches': missing_p,
            })

        # Prefer linked user's department, fall back to device department
        linked = getattr(emp, 'linked_user', None)
        dept = ''
        if linked and linked.department:
            dept = linked.department.name
        elif emp.department:
            dept = emp.department

        # Shift info
        shift_info = None
        if shift:
            shift_info = {
                'id': shift.id,
                'name': shift.name,
                'type': shift.shift_type,
                'slots': [(label, t.strftime('%H:%M')) for label, t in shift.time_slots],
            }

        # Calculate penalties
        late_penalty = _calc_penalty(shift_penalties, 'late', total_late)
        early_penalty = _calc_penalty(shift_penalties, 'early', total_early)

        rows.append({
            'user_id': emp_uid,
            'username': linked.username if linked else '',
            'employee_code': emp.employee_code or '',
            'employee_name': emp.display_name,
            'department': dept,
            'shift': shift_info,
            'summary': {
                'total_days': len(days),
                'present': len(days) - total_absent,
                'absent': total_absent,
                'late': total_late,
                'early_leave': total_early,
                'late_minutes': total_late_minutes,
                'early_minutes': total_early_minutes,
                'ot_hours': round(total_ot_minutes / 60, 1),
                'work_hours': round(total_work_minutes / 60, 1),
                'late_penalty': late_penalty,
                'early_penalty': early_penalty,
                'total_penalty': late_penalty + early_penalty,
            },
            'daily': daily,
        })

    return rows, days


class AttendanceReportView(APIView):
    """Monthly/daily attendance report with summary."""

    def get(self, request):
        user = request.user
        allowed = get_allowed_pages(user)
        # This view serves both 'report' and 'monthly' pages
        page_key = request.query_params.get('_page', 'report')
        if page_key not in allowed:
            return Response({'detail': 'Bạn không có quyền truy cập.'}, status=status.HTTP_403_FORBIDDEN)

        date_from_str = request.query_params.get('date_from')
        date_to_str = request.query_params.get('date_to')
        user_id = request.query_params.get('user_id')

        # Non-admin without view_all → force own data only
        if not can_view_all_on_page(user, page_key):
            att_emp = getattr(user, 'attendance_employee', None)
            if att_emp:
                user_id = att_emp.user_id
            else:
                # No mapping → return empty
                return Response({
                    'date_from': '', 'date_to': '', 'work_start': WORK_START.strftime('%H:%M'),
                    'work_end': WORK_END.strftime('%H:%M'), 'total_employees': 0, 'employees': [],
                })

        if not date_from_str or not date_to_str:
            # Default: current month
            today = timezone.now().date()
            date_from = today.replace(day=1)
            date_to = today
        else:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Định dạng ngày không hợp lệ (YYYY-MM-DD)'},
                                status=status.HTTP_400_BAD_REQUEST)

        rows, days = _build_report_data(date_from, date_to, user_id)

        return Response({
            'date_from': date_from.strftime('%Y-%m-%d'),
            'date_to': date_to.strftime('%Y-%m-%d'),
            'work_start': WORK_START.strftime('%H:%M'),
            'work_end': WORK_END.strftime('%H:%M'),
            'total_employees': len(rows),
            'employees': rows,
        })


class AttendanceReportExportView(APIView):
    """Export attendance report to Excel."""

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        user = request.user
        allowed = get_allowed_pages(user)
        page_key = request.query_params.get('_page', 'report')
        if page_key not in allowed:
            return Response({'detail': 'Bạn không có quyền truy cập.'}, status=status.HTTP_403_FORBIDDEN)

        date_from_str = request.query_params.get('date_from')
        date_to_str = request.query_params.get('date_to')
        user_id = request.query_params.get('user_id')

        if not can_view_all_on_page(user, page_key):
            att_emp = getattr(user, 'attendance_employee', None)
            if att_emp:
                user_id = att_emp.user_id
            else:
                user_id = '__none__'  # yields empty result

        if not date_from_str or not date_to_str:
            today = timezone.now().date()
            date_from = today.replace(day=1)
            date_to = today
        else:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Định dạng ngày không hợp lệ'},
                                status=status.HTTP_400_BAD_REQUEST)

        rows, days = _build_report_data(date_from, date_to, user_id)

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1677ff')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # ── Sheet 0: Bảng công ngang (calendar grid) ── only for monthly
        if page_key == 'monthly':
            ws_grid = wb.active
            ws_grid.title = 'Bảng công'

            present_fill = PatternFill('solid', fgColor='F6FFED')
            absent_fill_grid = PatternFill('solid', fgColor='FFF1F0')
            late_fill_grid = PatternFill('solid', fgColor='FFF7E6')
            weekend_fill = PatternFill('solid', fgColor='F5F5F5')
            header_fill_grid = PatternFill('solid', fgColor='1677ff')
            summary_fill = PatternFill('solid', fgColor='E6F7FF')

            day_labels = ['CN', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7']

            num_days = len(days)
            total_cols = 2 + num_days + 2  # STT, Họ tên, days..., Đi làm, Vắng

            # Title row
            ws_grid.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            title_cell_g = ws_grid.cell(row=1, column=1,
                value=f'BẢNG CHẤM CÔNG THÁNG {date_from.strftime("%m/%Y")}')
            title_cell_g.font = Font(bold=True, size=14)
            title_cell_g.alignment = Alignment(horizontal='center')

            # Header row 1: STT, Họ tên, day-of-week labels, Đi làm, Vắng
            # Header row 2: (merged), (merged), day numbers, (merged), (merged)
            ws_grid.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
            stt_cell = ws_grid.cell(row=3, column=1, value='STT')
            stt_cell.font = header_font
            stt_cell.fill = header_fill_grid
            stt_cell.alignment = header_align
            stt_cell.border = thin_border
            ws_grid.cell(row=4, column=1).border = thin_border
            ws_grid.cell(row=4, column=1).fill = header_fill_grid

            ws_grid.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
            name_cell = ws_grid.cell(row=3, column=2, value='Họ tên')
            name_cell.font = header_font
            name_cell.fill = header_fill_grid
            name_cell.alignment = header_align
            name_cell.border = thin_border
            ws_grid.cell(row=4, column=2).border = thin_border
            ws_grid.cell(row=4, column=2).fill = header_fill_grid

            from datetime import date as date_cls
            for i, day in enumerate(days):
                col = 3 + i
                dow = day.isoweekday() % 7  # 0=Sun
                is_weekend = dow == 0 or dow == 6

                # Row 3: day-of-week
                c1 = ws_grid.cell(row=3, column=col, value=day_labels[dow])
                c1.font = Font(bold=True, color='FFFFFF' if not is_weekend else 'FF4D4F', size=10)
                c1.fill = header_fill_grid if not is_weekend else PatternFill('solid', fgColor='003A8C')
                c1.alignment = header_align
                c1.border = thin_border

                # Row 4: day number
                c2 = ws_grid.cell(row=4, column=col, value=day.day)
                c2.font = Font(bold=True, color='FFFFFF' if not is_weekend else 'FF4D4F', size=10)
                c2.fill = header_fill_grid if not is_weekend else PatternFill('solid', fgColor='003A8C')
                c2.alignment = header_align
                c2.border = thin_border

            # Đi làm, Vắng header (merged rows 3-4)
            for offset, label in [(0, 'Đi làm'), (1, 'Vắng')]:
                col = 3 + num_days + offset
                ws_grid.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
                c = ws_grid.cell(row=3, column=col, value=label)
                c.font = header_font
                c.fill = header_fill_grid
                c.alignment = header_align
                c.border = thin_border
                ws_grid.cell(row=4, column=col).border = thin_border
                ws_grid.cell(row=4, column=col).fill = header_fill_grid

            # Data rows
            for idx, row_data in enumerate(rows, 1):
                r = 4 + idx
                day_map = {}
                for d in row_data['daily']:
                    day_map[d['date']] = d

                # STT
                ws_grid.cell(row=r, column=1, value=idx).border = thin_border

                # Name + info
                name_parts = [row_data['employee_name']]
                if row_data.get('department'):
                    name_parts.append(row_data['department'])
                if row_data.get('shift'):
                    name_parts.append(row_data['shift']['name'])
                # Show system username if linked, otherwise machine user_id
                display_id = row_data['username'] if row_data.get('username') else row_data['user_id']
                name_parts.append(f"TK: {display_id}")
                nc = ws_grid.cell(row=r, column=2, value='\n'.join(name_parts))
                nc.alignment = Alignment(vertical='center', wrap_text=True)
                nc.border = thin_border

                for i, day in enumerate(days):
                    col = 3 + i
                    date_str = day.strftime('%Y-%m-%d')
                    dow = day.isoweekday() % 7
                    is_weekend = dow == 0 or dow == 6
                    info = day_map.get(date_str)

                    cell = ws_grid.cell(row=r, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(size=9)

                    if is_weekend and (not info or info['status'] == 'absent'):
                        cell.value = '—'
                        cell.fill = weekend_fill
                        cell.font = Font(size=9, color='BBBBBB')
                    elif not info or info['status'] == 'absent':
                        cell.value = 'Vắng'
                        cell.fill = absent_fill_grid
                        cell.font = Font(size=9, color='FF4D4F', bold=True)
                    else:
                        # Show check-in / check-out
                        ci = (info['check_in'] or '—')[:5]
                        co = (info['check_out'] or '—')[:5]
                        cell.value = f"{ci}\n{co}"
                        if 'late' in info['status']:
                            cell.fill = late_fill_grid
                            cell.font = Font(size=9, color='D46B08')
                        else:
                            cell.fill = present_fill

                # Summary cols
                s = row_data['summary']
                pc = ws_grid.cell(row=r, column=3 + num_days, value=s['present'])
                pc.border = thin_border
                pc.alignment = Alignment(horizontal='center', vertical='center')
                pc.fill = summary_fill
                pc.font = Font(bold=True, color='52C41A')

                ac = ws_grid.cell(row=r, column=3 + num_days + 1, value=s['absent'])
                ac.border = thin_border
                ac.alignment = Alignment(horizontal='center', vertical='center')
                ac.fill = summary_fill
                ac.font = Font(bold=True, color='FF4D4F')

            # Column widths
            ws_grid.column_dimensions['A'].width = 5
            ws_grid.column_dimensions['B'].width = 22
            for i in range(num_days):
                col_letter = openpyxl.utils.get_column_letter(3 + i)
                ws_grid.column_dimensions[col_letter].width = 7
            ws_grid.column_dimensions[openpyxl.utils.get_column_letter(3 + num_days)].width = 8
            ws_grid.column_dimensions[openpyxl.utils.get_column_letter(3 + num_days + 1)].width = 7

            # Row height for data rows
            for r in range(5, 5 + len(rows)):
                ws_grid.row_dimensions[r].height = 35

        # ── Sheet 1: Summary ──
        if page_key == 'monthly':
            ws = wb.create_sheet('Tổng hợp')
        else:
            ws = wb.active
            ws.title = 'Tổng hợp'

        # Title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f'BÁO CÁO CHẤM CÔNG: {date_from.strftime("%d/%m/%Y")} – {date_to.strftime("%d/%m/%Y")}'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:I2')
        ws['A2'].value = f'Ca làm việc: {WORK_START.strftime("%H:%M")} – {WORK_END.strftime("%H:%M")}'
        ws['A2'].alignment = Alignment(horizontal='center')

        headers = ['STT', 'Mã NV', 'Họ tên', 'Phòng ban', 'Số ngày đi làm',
                   'Đi muộn', 'Về sớm', 'Vắng mặt', 'Tăng ca (giờ)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for idx, row in enumerate(rows, 1):
            s = row['summary']
            display_id = row['username'] if row.get('username') else row['user_id']
            values = [idx, display_id, row['employee_name'], row['department'],
                      s['present'], s['late'], s['early_leave'], s['absent'], s['ot_hours']]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=4 + idx, column=col, value=val)
                cell.border = thin_border
                if col >= 5:
                    cell.alignment = Alignment(horizontal='center')

        # Auto-width
        for col_cells in ws.columns:
            real_cells = [c for c in col_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)]
            if not real_cells:
                continue
            max_len = max((len(str(c.value or '')) for c in real_cells), default=10)
            ws.column_dimensions[real_cells[0].column_letter].width = min(max_len + 3, 30)

        # ── Sheet 2: Daily detail ──
        ws2 = wb.create_sheet('Chi tiết ngày')

        headers2 = ['Mã NV', 'Họ tên', 'Ngày', 'Check-in', 'Check-out',
                     'Trạng thái', 'Muộn (phút)', 'Sớm (phút)', 'OT (phút)', 'Làm việc (phút)']
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        status_vn = {
            'present': 'Đúng giờ', 'late': 'Đi muộn', 'early': 'Về sớm',
            'late+early': 'Muộn+Sớm', 'absent': 'Vắng mặt',
        }
        absent_fill = PatternFill('solid', fgColor='FFF0F0')
        late_fill = PatternFill('solid', fgColor='FFF8E1')

        r = 2
        for row in rows:
            display_id = row['username'] if row.get('username') else row['user_id']
            for d in row['daily']:
                values = [
                    display_id, row['employee_name'], d['date'],
                    d['check_in'] or '', d['check_out'] or '',
                    status_vn.get(d['status'], d['status']),
                    d['late_minutes'], d['early_minutes'], d['ot_minutes'], d['work_minutes'],
                ]
                for col, val in enumerate(values, 1):
                    cell = ws2.cell(row=r, column=col, value=val)
                    cell.border = thin_border
                    if d['status'] == 'absent':
                        cell.fill = absent_fill
                    elif 'late' in d['status']:
                        cell.fill = late_fill
                r += 1

        for col in ws2.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)

        # ── Sheet 3: Đi muộn / Về sớm ──
        ws3 = wb.create_sheet('Đi muộn - Về sớm')
        h3 = ['Mã NV', 'Họ tên', 'Phòng ban', 'Ca', 'Ngày', 'Loại vi phạm',
               'Mốc', 'Số phút', 'Check-in', 'Check-out']
        for col, h in enumerate(h3, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        r3 = 2
        for row in rows:
            display_id = row['username'] if row.get('username') else row['user_id']
            for d in row['daily']:
                if d['late_minutes'] > 0:
                    vals = [display_id, row['employee_name'], row['department'],
                            row['shift']['name'] if row['shift'] else '',
                            d['date'], 'Đi muộn', d.get('late_label') or '',
                            d['late_minutes'], d['check_in'] or '', d['check_out'] or '']
                    for col, val in enumerate(vals, 1):
                        cell = ws3.cell(row=r3, column=col, value=val)
                        cell.border = thin_border
                        cell.fill = late_fill
                    r3 += 1
                if d['early_minutes'] > 0:
                    vals = [display_id, row['employee_name'], row['department'],
                            row['shift']['name'] if row['shift'] else '',
                            d['date'], 'Về sớm', d.get('early_label') or '',
                            d['early_minutes'], d['check_in'] or '', d['check_out'] or '']
                    for col, val in enumerate(vals, 1):
                        cell = ws3.cell(row=r3, column=col, value=val)
                        cell.border = thin_border
                        cell.fill = PatternFill('solid', fgColor='E6F7FF')
                    r3 += 1

        # Summary row with penalties
        r3 += 1
        ws3.cell(row=r3, column=1, value='TỔNG HỢP PHẠT').font = Font(bold=True, size=12)
        r3 += 1
        penalty_headers = ['Mã NV', 'Họ tên', 'Số lần muộn', 'Tổng phút muộn',
                           'Phạt muộn (VNĐ)', 'Số lần về sớm', 'Tổng phút về sớm',
                           'Phạt về sớm (VNĐ)', 'Tổng phạt (VNĐ)']
        for col, h in enumerate(penalty_headers, 1):
            cell = ws3.cell(row=r3, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='722ed1')
            cell.alignment = header_align
            cell.border = thin_border
        r3 += 1
        for row in rows:
            s = row['summary']
            display_id = row['username'] if row.get('username') else row['user_id']
            if s['late'] > 0 or s['early_leave'] > 0:
                vals = [display_id, row['employee_name'],
                        s['late'], s.get('late_minutes', 0), s.get('late_penalty', 0),
                        s['early_leave'], s.get('early_minutes', 0), s.get('early_penalty', 0),
                        s.get('total_penalty', 0)]
                for col, val in enumerate(vals, 1):
                    cell = ws3.cell(row=r3, column=col, value=val)
                    cell.border = thin_border
                    if col >= 3:
                        cell.alignment = Alignment(horizontal='center')
                r3 += 1

        for col_cells in ws3.columns:
            real_cells = [c for c in col_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)]
            if not real_cells:
                continue
            max_len = max((len(str(c.value or '')) for c in real_cells), default=10)
            ws3.column_dimensions[real_cells[0].column_letter].width = min(max_len + 3, 30)

        # ── Sheet 4: Thiếu chấm công ──
        ws4 = wb.create_sheet('Thiếu chấm công')
        h4 = ['Mã NV', 'Họ tên', 'Phòng ban', 'Ca', 'Ngày', 'Lần chấm thiếu', 'Check-in', 'Check-out', 'Số lần chấm']
        for col, h in enumerate(h4, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='fa8c16')
            cell.alignment = header_align
            cell.border = thin_border

        r4 = 2
        for row in rows:
            display_id = row['username'] if row.get('username') else row['user_id']
            for d in row['daily']:
                mp = d.get('missing_punches', [])
                if mp:
                    vals = [display_id, row['employee_name'], row['department'],
                            row['shift']['name'] if row['shift'] else '',
                            d['date'], ', '.join(mp),
                            d['check_in'] or '', d['check_out'] or '',
                            len(d.get('punches', []))]
                    for col, val in enumerate(vals, 1):
                        cell = ws4.cell(row=r4, column=col, value=val)
                        cell.border = thin_border
                        if d['status'] == 'absent':
                            cell.fill = absent_fill
                    r4 += 1

        for col_cells in ws4.columns:
            real_cells = [c for c in col_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)]
            if not real_cells:
                continue
            max_len = max((len(str(c.value or '')) for c in real_cells), default=10)
            ws4.column_dimensions[real_cells[0].column_letter].width = min(max_len + 3, 30)

        # Write to response
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'BaoCaoChamCong_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.xlsx'
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ─── Attendance Permission Management (admin-only) ───────────────────

class AttendancePermissionListCreateView(APIView):
    """List / create attendance permissions. Admin only."""
    permission_classes = [IsAdmin]

    def get(self, request):
        qs = AttendancePermission.objects.select_related('user', 'department').order_by('-created_at')
        serializer = AttendancePermissionSerializer(qs, many=True)
        return Response({'results': serializer.data})

    def post(self, request):
        serializer = AttendancePermissionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class AttendancePermissionDeleteView(APIView):
    """Delete an attendance permission. Admin only."""
    permission_classes = [IsAdmin]

    def delete(self, request, pk):
        try:
            obj = AttendancePermission.objects.get(pk=pk)
        except AttendancePermission.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MyAttendanceInfoView(APIView):
    """Return the current user's attendance mapping & allowed pages."""

    def get(self, request):
        user = request.user
        allowed = get_allowed_pages(user)
        att_emp = user.attendance_employee

        # Determine which pages have can_view_all
        view_all_pages = set()
        for page in allowed:
            if can_view_all_on_page(user, page):
                view_all_pages.add(page)

        return Response({
            'attendance_employee_id': att_emp.user_id if att_emp else None,
            'attendance_employee_name': att_emp.name if att_emp else None,
            'allowed_pages': sorted(allowed),
            'view_all_pages': sorted(view_all_pages),
        })


class UserAttendanceEmployeeMappingView(APIView):
    """Admin: map a User to an attendance.Employee by user_id on device."""
    permission_classes = [IsAdmin]

    def post(self, request):
        from users.models import User
        erp_user_id = request.data.get('user_id')
        device_employee_id = request.data.get('attendance_employee_id')  # user_id on device

        if not erp_user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            erp_user = User.objects.get(pk=erp_user_id)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

        if device_employee_id:
            emp = Employee.objects.filter(user_id=str(device_employee_id)).first()
            if not emp:
                return Response({'error': f'Không tìm thấy nhân viên chấm công với mã {device_employee_id}'},
                                status=status.HTTP_404_NOT_FOUND)
            # If another user is already linked to this employee, unlink them first
            if hasattr(emp, 'linked_user') and emp.linked_user and emp.linked_user.pk != erp_user.pk:
                old_user = emp.linked_user
                old_user.attendance_employee = None
                old_user.save(update_fields=['attendance_employee'])
            erp_user.attendance_employee = emp
        else:
            erp_user.attendance_employee = None

        erp_user.save(update_fields=['attendance_employee'])
        return Response({
            'user_id': erp_user.pk,
            'username': erp_user.username,
            'attendance_employee_id': erp_user.attendance_employee.user_id if erp_user.attendance_employee else None,
            'attendance_employee_name': erp_user.attendance_employee.name if erp_user.attendance_employee else None,
        })


class EmployeeToggleActiveView(APIView):
    """Admin: toggle is_active on an attendance Employee."""
    permission_classes = [IsAdmin]

    def patch(self, request, pk):
        try:
            emp = Employee.objects.get(pk=pk)
        except Employee.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        is_active = request.data.get('is_active')
        if is_active is None:
            return Response({'error': 'is_active is required'}, status=status.HTTP_400_BAD_REQUEST)

        emp.is_active = bool(is_active)
        emp.save(update_fields=['is_active'])
        return Response({'id': emp.pk, 'is_active': emp.is_active})


class EmployeeBulkToggleActiveView(APIView):
    """Admin: bulk set is_active for multiple employees."""
    permission_classes = [IsAdmin]

    def patch(self, request):
        ids = request.data.get('ids', [])
        is_active = request.data.get('is_active')
        if not ids or is_active is None:
            return Response({'error': 'ids and is_active are required'}, status=status.HTTP_400_BAD_REQUEST)
        updated = Employee.objects.filter(pk__in=ids).update(is_active=bool(is_active))
        return Response({'updated': updated, 'is_active': bool(is_active)})


# ─── Work Shift management ──────────────────────────────────

class WorkShiftListCreateView(APIView):
    """List / create work shifts."""
    permission_classes = [IsAdmin]

    def get(self, request):
        shifts = WorkShift.objects.annotate(employee_count=Count('employees')).all()
        data = WorkShiftSerializer(shifts, many=True).data
        # Attach employee_count from annotation
        shift_list = list(shifts)
        for i, d in enumerate(data):
            d['employee_count'] = shift_list[i].employee_count
        return Response(data)

    def post(self, request):
        serializer = WorkShiftSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class WorkShiftDetailView(APIView):
    """Retrieve / update / delete a work shift."""
    permission_classes = [IsAdmin]

    def get_object(self, pk):
        try:
            return WorkShift.objects.get(pk=pk)
        except WorkShift.DoesNotExist:
            return None

    def get(self, request, pk):
        obj = self.get_object(pk)
        if not obj:
            return Response(status=status.HTTP_404_NOT_FOUND)
        return Response(WorkShiftSerializer(obj).data)

    def put(self, request, pk):
        obj = self.get_object(pk)
        if not obj:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = WorkShiftSerializer(obj, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        obj = self.get_object(pk)
        if not obj:
            return Response(status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class EmployeeAssignShiftView(APIView):
    """Assign a shift to one or multiple employees."""
    permission_classes = [IsAdmin]

    def patch(self, request):
        ids = request.data.get('ids', [])
        shift_id = request.data.get('shift_id')  # null to clear
        if not ids:
            return Response({'error': 'ids is required'}, status=status.HTTP_400_BAD_REQUEST)
        if shift_id is not None:
            try:
                WorkShift.objects.get(pk=shift_id)
            except WorkShift.DoesNotExist:
                return Response({'error': 'Ca không tồn tại'}, status=status.HTTP_404_NOT_FOUND)
        updated = Employee.objects.filter(pk__in=ids).update(shift_id=shift_id)
        return Response({'updated': updated})


# ─── Late/Early Rules & Penalty Config ───────────────────

class LateEarlyRuleListCreateView(APIView):
    """List/create late-early rules for a shift."""
    permission_classes = [IsAdmin]

    def get(self, request):
        shift_id = request.query_params.get('shift_id')
        qs = LateEarlyRule.objects.select_related('shift').all()
        if shift_id:
            qs = qs.filter(shift_id=shift_id)
        return Response(LateEarlyRuleSerializer(qs, many=True).data)

    def post(self, request):
        serializer = LateEarlyRuleSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class LateEarlyRuleDetailView(APIView):
    permission_classes = [IsAdmin]

    def put(self, request, pk):
        try:
            obj = LateEarlyRule.objects.get(pk=pk)
        except LateEarlyRule.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = LateEarlyRuleSerializer(obj, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        try:
            obj = LateEarlyRule.objects.get(pk=pk)
        except LateEarlyRule.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class LateEarlyRuleBulkView(APIView):
    """Bulk save rules for a shift (replace all)."""
    permission_classes = [IsAdmin]

    def post(self, request):
        shift_id = request.data.get('shift_id')
        rules = request.data.get('rules', [])
        if not shift_id:
            return Response({'error': 'shift_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            shift = WorkShift.objects.get(pk=shift_id)
        except WorkShift.DoesNotExist:
            return Response({'error': 'Ca không tồn tại'}, status=status.HTTP_404_NOT_FOUND)
        # Delete existing rules for this shift and recreate
        LateEarlyRule.objects.filter(shift=shift).delete()
        created = []
        for r in rules:
            obj = LateEarlyRule.objects.create(
                shift=shift,
                rule_type=r['rule_type'],
                from_minutes=r['from_minutes'],
                to_minutes=r.get('to_minutes'),
                label=r.get('label', ''),
            )
            created.append(obj)
        return Response(LateEarlyRuleSerializer(created, many=True).data)


class PenaltyConfigListCreateView(APIView):
    """List/create penalty configs for a shift."""
    permission_classes = [IsAdmin]

    def get(self, request):
        shift_id = request.query_params.get('shift_id')
        qs = PenaltyConfig.objects.select_related('shift').all()
        if shift_id:
            qs = qs.filter(shift_id=shift_id)
        return Response(PenaltyConfigSerializer(qs, many=True).data)

    def post(self, request):
        serializer = PenaltyConfigSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class PenaltyConfigDetailView(APIView):
    permission_classes = [IsAdmin]

    def put(self, request, pk):
        try:
            obj = PenaltyConfig.objects.get(pk=pk)
        except PenaltyConfig.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = PenaltyConfigSerializer(obj, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, pk):
        try:
            obj = PenaltyConfig.objects.get(pk=pk)
        except PenaltyConfig.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class PenaltyConfigBulkView(APIView):
    """Bulk save penalty configs for a shift (replace all)."""
    permission_classes = [IsAdmin]

    def post(self, request):
        shift_id = request.data.get('shift_id')
        configs = request.data.get('configs', [])
        if not shift_id:
            return Response({'error': 'shift_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            shift = WorkShift.objects.get(pk=shift_id)
        except WorkShift.DoesNotExist:
            return Response({'error': 'Ca không tồn tại'}, status=status.HTTP_404_NOT_FOUND)
        PenaltyConfig.objects.filter(shift=shift).delete()
        created = []
        for c in configs:
            obj = PenaltyConfig.objects.create(
                shift=shift,
                rule_type=c['rule_type'],
                from_count=c['from_count'],
                to_count=c.get('to_count'),
                penalty_amount=c.get('penalty_amount', 0),
            )
            created.append(obj)
        return Response(PenaltyConfigSerializer(created, many=True).data)
